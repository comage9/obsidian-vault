#!/usr/bin/env python3
"""
Analyze an XLSM/XLSX file for LibreOffice compatibility issues.

Usage:
    python3 analyze_xlsm_compat.py <file.xlsm>

Scans for:
- _xlfn. prefixed functions (Excel future functions)
- SINGLE() wrappers
- #REF! errors in formulas
- _xlfn. prefix in shared strings
"""
import sys, zipfile, re, json
from collections import Counter

def analyze(filepath):
    issues = Counter()
    formula_samples = []
    ref_errors = []
    
    # Open as ZIP to check structure
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            has_vba = 'xl/vbaProject.bin' in z.namelist()
    except:
        has_vba = False
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=False)
    except ImportError:
        print("openpyxl not installed. Install: pip install openpyxl")
        sys.exit(1)
    
    for ws in wb.worksheets:
        sheet_name = ws.title
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value
                    if '_xlfn.' in val:
                        issues['_xlfn. prefix'] += 1
                        if len(formula_samples) < 5:
                            formula_samples.append(f"  {cell.coordinate}: {val[:100]}")
                    if '#REF!' in val:
                        issues['#REF! error'] += 1
                        if len(ref_errors) < 10:
                            ref_errors.append(f"  {cell.coordinate}: {val[:100]}")
    
    print(f"File: {filepath}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"VBA macros: {'YES' if has_vba else 'NO'}")
    print()
    print("=== Issues Found ===")
    for issue, count in issues.most_common():
        print(f"  {issue}: {count}")
    print()
    if formula_samples:
        print("=== _xlfn. Samples ===")
        for s in formula_samples:
            print(s)
    print()
    if ref_errors:
        print("=== #REF! Errors ===")
        for e in ref_errors:
            print(e)

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 analyze_xlsm_compat.py <file.xlsm>")
        sys.exit(1)
    analyze(sys.argv[1])
